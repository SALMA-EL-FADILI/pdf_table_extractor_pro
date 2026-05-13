import logging
from typing import List, Dict, Optional

import pandas as pd

from .config import HAS_OPENPYXL, SECTION_DISPLAY_NAMES, COLORS, _TOTAL_RE
from .models import ParsedSection, SectionScore
from .pipeline import post_clean

log = logging.getLogger(__name__)


def _thin():
    from openpyxl.styles import Border, Side
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)


def export_to_excel(sections: List[ParsedSection],
                    out: str,
                    exercise_label: str = "N/A",
                    validation_scores: Optional[Dict[str, SectionScore]] = None,
                    ai_metadata: Optional[Dict[str, Dict]] = None) -> str:
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl non installé")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook(); wb.remove(wb.active)
    nf = '#,##0.00'; used_names: Dict[str, int] = {}

    if validation_scores:
        ws_q = wb.create_sheet(title=" Qualité Extraction")
        headers = ["Section", "Score", "Anomalies", "Statut IA", "Évaluation IA"]
        for ci, h in enumerate(headers, 1):
            c = ws_q.cell(1, ci, h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor=COLORS["hbg"])
            c.alignment = Alignment(horizontal="center")
        for ri, (sk, score_obj) in enumerate(validation_scores.items(), 2):
            ai_meta  = (ai_metadata or {}).get(sk, {})
            row_data = [
                SECTION_DISPLAY_NAMES.get(sk, sk),
                f"{score_obj.score:.0f}/100",
                len(score_obj.anomalies),
                " Raffiné" if ai_meta.get("corrections_count", 0) > 0 else (
                    " Analysé" if sk in (ai_metadata or {}) else "—"
                ),
                ai_meta.get("global_assessment", ""),
            ]
            for ci, val in enumerate(row_data, 1):
                c = ws_q.cell(ri, ci, val)
                if ci == 2:
                    score_num = score_obj.score
                    color = "2ECC71" if score_num >= 90 else "F39C12" if score_num >= 70 else "E74C3C"
                    c.fill = PatternFill("solid", fgColor=color)
                    c.font = Font(bold=True, color="FFFFFF")
        for ci in range(1, 6):
            ws_q.column_dimensions[get_column_letter(ci)].width = [28, 10, 12, 14, 50][ci-1]

    for sec in sections:
        df = post_clean(sec.dataframe)
        if df is None or df.empty: continue
        raw_name = sec.display_name[:28]
        cnt = used_names.get(raw_name, 0) + 1; used_names[raw_name] = cnt
        sheet_name = raw_name if cnt == 1 else f"{raw_name[:25]}_{cnt}"

        score_obj = (validation_scores or {}).get(sec.section_key)
        if score_obj:
            emoji = "" if score_obj.score >= 90 else "" if score_obj.score >= 70 else ""
            sheet_name = f"{emoji} {sheet_name}"[:31]

        ws = wb.create_sheet(title=sheet_name)
        nc = len(df.columns)

        title_text = sec.display_name
        if exercise_label and exercise_label != "N/A":
            title_text += f"  —  Exercice {exercise_label}"

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=nc)
        tc = ws.cell(1, 1, title_text)
        tc.font      = Font(name="Arial", bold=True, size=12, color=COLORS["hfg"])
        tc.fill      = PatternFill("solid", fgColor=COLORS["hbg"])
        tc.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22

        for ci, col in enumerate(df.columns, 1):
            c = ws.cell(2, ci, col)
            c.font      = Font(name="Arial", bold=True, size=10, color=COLORS["hfg"])
            c.fill      = PatternFill("solid", fgColor=COLORS["sbg"])
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border    = _thin()
        ws.row_dimensions[2].height = 32

        for ri, (_, row) in enumerate(df.iterrows(), 3):
            lbl      = str(row.get("Libellé", "")) if pd.notna(row.get("Libellé")) else ""
            is_total = bool(_TOTAL_RE.search(lbl))
            is_alt   = ri % 2 == 0
            for ci, col in enumerate(df.columns, 1):
                val = row[col]
                if pd.isna(val): val = None
                c = ws.cell(ri, ci, val); c.border = _thin()
                if is_total:
                    c.font = Font(name="Arial", bold=True, size=9)
                    c.fill = PatternFill("solid", fgColor=COLORS["tbg"])
                elif is_alt:
                    c.font = Font(name="Arial", size=9)
                    c.fill = PatternFill("solid", fgColor=COLORS["alt"])
                else:
                    c.font = Font(name="Arial", size=9)
                if col == "Libellé":
                    c.alignment = Alignment(horizontal="left", vertical="center",
                                            indent=1 if lbl.startswith("•") else 0)
                elif isinstance(val, (int, float)):
                    c.number_format = nf
                    c.alignment     = Alignment(horizontal="right", vertical="center")
                else:
                    c.alignment = Alignment(horizontal="center", vertical="center")

        for ci, col in enumerate(df.columns, 1):
            cl = get_column_letter(ci)
            if col == "Libellé":
                mx = max((len(str(v)) for v in df[col].dropna().head(100)), default=30)
                ws.column_dimensions[cl].width = min(65, max(32, mx + 2))
            else:
                ws.column_dimensions[cl].width = 20
        ws.freeze_panes = "B3"

    wb.save(out)
    log.info(f" Excel : {out}")
    return out
